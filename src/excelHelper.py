import openpyxl
from openpyxl.styles import PatternFill

def FillBest(worksheet,best_val):
    yellow = "00FFFF00"
    for row in worksheet.iter_rows():
        for cell in row:
            cellContent = str(cell.value)
            if cellContent == f"{best_val}":
                cell.fill = PatternFill(start_color=yellow, end_color=yellow,
                                        fill_type = "solid")


