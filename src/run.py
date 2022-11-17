import statistics
import openpyxl
import excelHelper
import functions
from enums import enumFunctions, enumCrossovers, enumSelections, enumDecreaseTemp
from algorithms import GeneticAlgorithm as GA, SimulatedAnnealing as SA, GreyWolfOptimization as GWO, HillClimbing as HC, HarmonySearchAlgorithm as HS


num_of_run = 2 # 5
dim =  10 # 30
results = []
arr_functions = enumFunctions.Functions
pop_sizes = [100] # [250, 500, 1000, 3000]
num_of_generations = [100] # [250, 500, 1000, 1500]



def run_ga():
    results.clear()
    func_name = None
    result = None
    best_val = None
    mut_probs = [0.01] # [0.01, 0.02, 0.05, 0.1, 0.15]
    crossover_types = enumCrossovers.Crossovers
    selection_types = enumSelections.Selections

    workbook = openpyxl.Workbook()
    workbook.remove(workbook.get_sheet_by_name('Sheet'))

    for func in arr_functions:
        func_name = arr_functions(func).name

        workbook.create_sheet(func_name)
        worksheet = workbook.get_sheet_by_name(func_name)
        worksheet.append(("Function", "Number of Generation", "Mutation Probability", "Crossover Type",
                    "Selection Type", "Best Fitness", "Best Average Fitness", "Std_Dev Fitness"))

        obj_func = functions.selectFunction(func)
        lb, ub = functions.getBoundaries(func, dim)

        for pop_size in pop_sizes:
            for num_of_gen in num_of_generations:
                for mut_prob in mut_probs:
                    for crossover_type in crossover_types:
                        for selection_type in selection_types:
                            for i in range(num_of_run):
                                result = GA.GA(
                                    obj_func, lb, ub, dim, pop_size, num_of_gen, mut_prob, crossover_type, selection_type)
                                results.append(result.best)

                            avg = statistics.mean(results)
                            std_dev = statistics.stdev(results)
                            worksheet.append((func_name, num_of_gen,
                                              mut_prob, crossover_types(
                                                  crossover_type).name,
                                              selection_types(selection_type).name, result.best, avg, std_dev))
                            workbook.save('ga_results.xlsx')

                            if best_val != None and best_val > avg:
                                best_val = avg
                            elif best_val == None:
                                best_val = avg

                            results.clear()

        excelHelper.FillBest(worksheet, best_val)
    workbook.save('ga_results.xlsx')


def run_sa():
    results.clear()
    func_name = None
    initial_temps = [1000, 5000, 10000]
    decrease_temp_types = enumDecreaseTemp.DecreaseTypes

    workbook = openpyxl.Workbook()
    workbook.remove(workbook.get_sheet_by_name('Sheet'))


    for func in arr_functions:
        func_name = arr_functions(func).name

        workbook.create_sheet(func_name)
        worksheet = workbook.get_sheet_by_name(func_name)
        worksheet.append(("Function", "Initial Temperature", "Decrease Temperature Type",
                    "Best Fitness", "Best Average Fitness", "Std_Dev Fitness"))

        obj_func = functions.selectFunction(func)
        lb, ub = functions.getBoundaries(func, dim)
        for initial_temp in initial_temps:
            for decrease_temp_type in decrease_temp_types:
                for i in range(num_of_run):
                    result = SA.simulated_annealing(min_values=[lb for _ in range(dim)],
                                                    max_values=[ub for _ in range(dim)], mu=0,
                                                    sigma=1, initial_temperature=initial_temp,
                                                    final_temperature=0.0001, alpha=0.1, target_function=obj_func, verbose=True, decrease_type=decrease_temp_type)
                    results.append(result)
                #!TO
                avg = statistics.mean(results)
                std_dev = statistics.stdev(results)

                worksheet.append((arr_functions(func).name, initial_temp,
                                  decrease_temp_type, result.best, avg, std_dev))
                workbook.save('sa_results.xlsx')

                if best_val != None and best_val > avg:
                    best_val = avg
                elif best_val == None:
                    best_val = avg

                results.clear()

        excelHelper.FillBest(worksheet, best_val)


def run_gwo():
    results.clear()
    func_name = None
    best_val = None

    workbook = openpyxl.Workbook()
    workbook.remove(workbook.get_sheet_by_name('Sheet'))


    for func in arr_functions:
        func_name = arr_functions(func).name

        workbook.create_sheet(func_name)
        worksheet = workbook.get_sheet_by_name(func_name)
        worksheet.append(("Function", "Population Size", "Number of Generations",
                "Best Fitness", "Best Average Fitness", "Std_Dev Fitness"))

        obj_func = functions.selectFunction(func)
        lb, ub = functions.getBoundaries(func,dim)
        for pop_size in pop_sizes:
            for num_of_gen in num_of_generations:
                for i in range(num_of_run):
                    result = GWO.GWO(obj_func, lb, ub,
                                     dim, pop_size, num_of_gen)
                    results.append(result.best)

                avg = statistics.mean(results)
                std_dev = statistics.stdev(results)

                worksheet.append((arr_functions(func).name, pop_size,
                                  num_of_gen, result.best, avg, std_dev))
                workbook.save('gwo_results.xlsx')

                if best_val != None and best_val > avg:
                    best_val = avg
                elif best_val == None:
                    best_val = avg

                results.clear()

        excelHelper.FillBest(worksheet, best_val)


def run_hc():
    num_of_generations = [250, 500, 1000, 1500, 2000, 5000]
    func_name = None
    results.clear()

    workbook = openpyxl.Workbook()
    workbook.remove(workbook.get_sheet_by_name('Sheet'))


    for func in arr_functions:
        func_name = arr_functions(func).name

        workbook.create_sheet(func_name)
        worksheet = workbook.get_sheet_by_name(func_name)
        worksheet.append(("Function", "Number of Generations",
                     "Best Fitness", "Best Average Fitness", "Std_Dev Fitness"))
            
        obj_func = functions.selectFunction(func)
        lb, ub = functions.getBoundaries(func.dim)
        for num_of_gen in num_of_generations:
            for i in range(num_of_run):
                pass # delete it
                #result = HC.HillClimbing(obj_func,lb,ub,dim,num_of_gen)
                # results.append(result.best)

                #avg     = statistics.mean(results)
                #std_dev = statistics.stdev(results)

                #worksheet.append((arr_functions(func).name, num_of_gen, result.best,avg,std_dev))
                # workbook.save('hc_results.xlsx')

                # if best_val != None and best_val > avg:
                #    best_val = avg
                # elif best_val == None:
                #    best_val = avg

                # results.clear()

        # excelHelper.FillBest(worksheet,best_val)


def run_hs():
    results.clear()
    func_name = None
    arr_hms = [250, 500, 1000, 5000]
    arr_bw = [0.1, 0.15, 0.2, 0.25, 0.30]
    arr_hmcr = [0.90, 0.92, 0.95, 0.97, 0.99]
    arr_par = [0.15, 0.2, 0.25, 0.3, 0.35, 0.4]

    workbook = openpyxl.Workbook()
    workbook.remove(workbook.get_sheet_by_name('Sheet'))


    for func in arr_functions:
        func_name = arr_functions(func).name

        workbook.create_sheet(func_name)
        worksheet = workbook.get_sheet_by_name(func_name)
        worksheet.append(("Function", "HMS", "BW", "HMCR", "PAR",
                     "Best Fitness", "Best Average Fitness", "Std_Dev Fitness"))

        obj_func = functions.selectFunction(func)
        lb,ub = functions.getBoundaries(func,dim)
        for hms in arr_hms:
            for bw in arr_bw:
                for hmcr in arr_hmcr:
                    for par in arr_par:
                        for i in range(num_of_run):
                            pass # delete it
                            #result = HS.HS(obj_func,hms,bw,hmcr,par)
                            # results.append(result.best)

                        #avg     = statistics.mean(results)
                        #std_dev = statistics.stdev(results)

                        #worksheet.append((arr_functions(func).name, num_of_gen, result.best,avg,std_dev))
                        # workbook.save('hs_results.xlsx')

                        # if best_val != None and best_val > avg:
                        #    best_val = avg
                        # elif best_val == None:
                        #    best_val = avg

                        # results.clear()

        # excelHelper.FillBest(worksheet,best_val)


def main():
    run_ga()
    # run_sa()
    # run_gwo()
    # run_hc()
    # run_hs()


if __name__ == "__main__":
    main()
